import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.styles import Font

def rowcol(driver, sheet):
    wait = WebDriverWait(driver, 3)
    time.sleep(0.5)
    # Find the number of rows and columns
    rows_elements1 = driver.find_elements(By.XPATH, '//table[@id="table-SHSummaryTable"]//tbody//tr')
    rows1 = len(rows_elements1)
    print(rows1)
    rows_elements2 = driver.find_elements(By.XPATH, '//table[@id="table-SHPublicShareHolderTable"]//tbody//tr')
    rows2 = len(rows_elements2)
    print(rows2)
    cols_elements = driver.find_elements(By.XPATH, '//table[@id="table-SHSummaryTable"]//tbody//tr[1]//td')  
    cols = len(cols_elements)
    print(cols)
    # Add a condition to catch if the driver acquired the incorrect number of rows and columns
    if rows1 == 1 or rows2 == 1 or cols == 1:
        print("problem; redoing yaah")
        time.sleep(0.25)
        rowcol(driver, sheet) # Repeat the function
    
    # Gather the data from the two tables after clicking to open them
    summ = wait.until(EC.element_to_be_clickable((By.ID, 'shpsummary')))
    summ.click()
    time.sleep(0.1)
    table1_data = [
        [driver.find_element(By.XPATH, f'//table[@id="table-SHSummaryTable"]//tbody//tr[{r+1}]//td[{c+1}]').text for c in range(cols)]
        for r in range(rows1)
    ]

    pubs = wait.until(EC.element_to_be_clickable((By.ID, 'shppublic')))
    pubs.click()
    time.sleep(0.1)
    table2_data = [
        [driver.find_element(By.XPATH, f'//table[@id="table-SHPublicShareHolderTable"]//tbody//tr[{r+1}]//td[{c+1}]').text for c in range(cols)]
        for r in range(rows2)
    ]

    # Write data to Excel in the desired order
    # Write first two rows of Summary Table
    current_row = 5
    for row_data in table1_data[:2]:
        for col_idx, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=current_row, column=col_idx).value = cell_value
            sheet.cell(row=current_row, column=col_idx).font = Font(bold=True)
        current_row += 1
    
    # Write all rows of Public Shareholders Table
    for row_data in table2_data:
        for col_idx, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=current_row, column=col_idx).value = cell_value
        current_row += 1
    
    # Write remaining rows of Summary Table
    for row_data in table1_data[2:]:
        for col_idx, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=current_row, column=col_idx).value = cell_value
            sheet.cell(row=current_row, column=col_idx).font = Font(bold=True)
        current_row += 1