import time
import os
import threading
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd


from selenium_utils import setup_driver, scroll
from excel_formatter import format_worksheet
from data_scraper import rowcol

def initiate(c_name):
    # Setup the driver
    driver, actions = setup_driver()
    driver.get("https://www.nseindia.com/companies-listing/corporate-filings-shareholding-pattern")  # Enter Website URL

    wait = WebDriverWait(driver, 3)
    time.sleep(1.5)
    input_field = wait.until(EC.element_to_be_clickable((By.XPATH, '//span[@class="twitter-typeahead"]//input[@placeholder="Company Name or Symbol"]')))
    input_field.click()
    # Enter the company name
    driver.implicitly_wait(1000)
    input_field.send_keys(c_name)

    # Choose the first autocorrect option
    search_result = wait.until(EC.element_to_be_clickable((By.XPATH, '//p[@class="line1"]')))
    search_valid = driver.find_element(By.XPATH, '//p[@class="line1"]//span')
    search_valid = search_valid.text.strip()

    # Add a condition to end the driver and return to the try again screen in case company name is invalid
    if search_valid == "No record found":
        driver.quit()
    search_result.click()
    time.sleep(0.5)

    # Search for the amount of rows within the main table of contents
    search_valid2 = driver.find_elements(By.XPATH, '//table[@id="shareHolderPattern"]//tbody//tr')
    search_valid2 = len(search_valid2)
    print(search_valid2)

    # If there is only one row, it means that there is no records available for that specific company
    if search_valid2 == 1:
        driver.quit()

    check_path = 'Shareholding_Patterns'
    if os.path.exists(check_path) and os.path.isdir(check_path):
        print("Folder exists")
    else:
        print("no folder")
        os.makedirs(check_path, exist_ok=True)

    # Initialise the excel file name and path
    xcel_path = f'Shareholding_Patterns/{search_valid}-SP.xlsx'
    print(xcel_path)

    # Find the total amount of shareholder patterns returned
    totalsize = driver.find_elements(By.XPATH, '//tbody//tr//td[@headers="companyN"]')
    print("found")
    size = 0
    flag = True
    size = len(totalsize)
    if os.path.isfile(xcel_path):
        # Set flag and load the already present file
        flag = False
        print(flag)
        wb_obj = openpyxl.load_workbook(xcel_path)
        
    else:
        print("new file")
        wb_obj = openpyxl.Workbook() # Open a new excel file for the company

    #size = 1
    print("size:", size)
    flagAlrUpd = True
    complocate = driver.find_element(By.XPATH, '//table[@id="shareHolderPattern"]//thead')
    driver.execute_script("arguments[0].scrollIntoView(true);", complocate)
    time.sleep(0.5)
    dates = []
    count = 0
    for i in range(size):  
        # Find the last 4 digits of the date and make sure to stop the process once files from 2015 and earlier start appearing
        date = driver.find_element(By.XPATH, f'//tbody//tr[{i+1}]//td[@headers="shareholdingOfTotalShares asOnDate"]')
        date = date.text.strip()
        final = date[-4:]
        if final == "2015":
            break
        print("\nposition:", i)
        try:
            sheet_date = wb_obj.sheetnames[count] # Gets the name of the first sheet in the excel file
            print(sheet_date)
            print("sheet position:", i)
            if date == sheet_date or dates[i] == dates[i-1]:
                print("Alr Present")
                count += 1
                continue
            if dates[i] != dates[i-1]:
                dates.append((i, date))
        except:
            print("hopefully not a true error")
            dates.append((i, date))
    for i, date in dates:
        print(f'{i} - {date}')
    print(len(dates))
    for i, date in dates:
        # Add a function to scroll to the element and make it visible therefore allowing the driver to click on the javascript button
        scroll(driver, i, i)
        flagAlrUpd = False

        # If it is the first iteration and the file didnt exist before, then choose the active sheet and edit its title
        if flag == True and i == 0:
            sheet = wb_obj.active
            sheet.title = date
        # Otherwise choose the appropriate sheet number and insert the date
        else:
            wb_obj.create_sheet(date, i)
            sheet = wb_obj[date]
        print(date)
        print("created sheet position:", i)
        
        # Acquire the number of rows and columns
        rowcol(driver, sheet)

        # Add the formatted headers for each sheet and save the excel then close the javascript
        format_worksheet(wb_obj, sheet)
        wb_obj.save(xcel_path)
        close = wait.until(EC.element_to_be_clickable((By.XPATH, '//div//button[@onclick="closeSHPopup()"]')))
        close.click()

    driver.quit()
    if flagAlrUpd == True:
        print("Already Updated")
    elif flag == True:
        print("Scraped")
    else:
        print("Updated")

if __name__ == "__main__":
    # df = pd.read_csv("ind_nifty50list.csv", usecols=["Symbol"])  # Load only the 'Symbol' column
    # symbols = df["Symbol"].tolist()
    # for symbol in symbols:
    #     print(symbol)
    #     initiate(symbol)
    initiate('CIPLA')