import openpyxl
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment, Border, Side

def format_worksheet(wb_obj, sheet):
    sheet.row_dimensions[1].height = 250
    sheet.row_dimensions[2].height = 30
    sheet.row_dimensions[3].height = 80
    sheet.row_dimensions[4].height = 50

    sheet.merge_cells('I1:L1')
    sheet.merge_cells('I2:L2')
    sheet.merge_cells('I3:K3')
    sheet.merge_cells('O1:P1')
    sheet.merge_cells('O2:P2')
    sheet.merge_cells('Q1:R1')
    sheet.merge_cells('Q2:R2')
    sheet.merge_cells('T1:V1')
    sheet.merge_cells('T2:V2')
    sheet.merge_cells('T3:V3')

    sheet.column_dimensions['B'].width = 35
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['I'].width = 15
    sheet.column_dimensions['K'].width = 15
    sheet.column_dimensions['N'].width = 15
    sheet.column_dimensions['S'].width = 15

    sheet.cell(row=1, column=1).value = "CATEGORY"
    sheet.cell(row=1, column=2).value = "CATEGORY OF SHAREHOLDER"
    sheet.cell(row=1, column=3).value = "NOS. OF SHAREHOLDERS"
    sheet.cell(row=1, column=4).value = "NO. OF FULLY PAID UP EQUITY SHARES HELD"
    sheet.cell(row=1, column=5).value = "NO. OF PARTLY PAID UP EQUITY SHARES HELD"
    sheet.cell(row=1, column=6).value = "NO. OF SHARES UNDERLYING DEPOSITORY RECIEPTS"
    sheet.cell(row=1, column=7).value = "TOTAL NOS. SHARES HELD"
    sheet.cell(row=1, column=8).value = "SHAREHOLDING AS A % OF TOTAL NO. OF SHARES (CALCULATED AS PER SCRR, 1957) AS A % OF (A+B+C2)"
    sheet.cell(row=1, column=9).value = "NUMBER OF VOTING RIGHTS HELD IN EACH CLASS OF SECURITIES"
    sheet.cell(row=1, column=13).value = "NO. OF SHARES UNDERLYING OUTSTANDING CONVERTIBLE SECURITIES (INCLUDING WARRANTS)"
    sheet.cell(row=1, column=14).value = "SHAREHOLDING , AS A % ASSUMING FULL CONVERSION OF CONVERTIBLE SECURITIES ( AS A PERCENTAGE OF DILUTED SHARE CAPITAL) AS A % OF (A+B+C2)"
    sheet.cell(row=1, column=15).value = "NUMBER OF LOCKED IN SHARES"
    sheet.cell(row=1, column=17).value = "NUMBER OF SHARES PLEDGED OR OTHERWISE ENCUMBERED"
    sheet.cell(row=1, column=19).value = "NUMBER OF EQUITY SHARES HELD IN DEMATERIALIZED FORM"
    sheet.cell(row=1, column=20).value = "SUB-CATEGORIZATION OF SHARES"
    sheet.cell(row=2, column=1).value = "(I)"
    sheet.cell(row=2, column=2).value = "(II)"
    sheet.cell(row=2, column=3).value = "(III)"
    sheet.cell(row=2, column=4).value = "(IV)"
    sheet.cell(row=2, column=5).value = "(V)"
    sheet.cell(row=2, column=6).value = "(VI)"
    sheet.cell(row=2, column=7).value = "(VII) = (IV) + (V) + (VI)"
    sheet.cell(row=2, column=8).value = "(VIII)"
    sheet.cell(row=2, column=9).value = "(IX)"
    sheet.cell(row=2, column=13).value = "(X)"
    sheet.cell(row=2, column=14).value = "(XI) = (VII) + (X)"
    sheet.cell(row=2, column=15).value = "(XII)"
    sheet.cell(row=2, column=17).value = "(XIII)"
    sheet.cell(row=2, column=19).value = "(XIV)"
    sheet.cell(row=2, column=20).value = "(XV)"
    sheet.cell(row=3, column=9).value = "NO. OF VOTING RIGHTS"
    sheet.cell(row=3, column=12).value = "TOTAL AS A % OF (A + B + C)"
    sheet.cell(row=3, column=15).value = "NO."
    sheet.cell(row=3, column=16).value = "AS A % OF TOTAL SHARES HELD"
    sheet.cell(row=3, column=17).value = "NO."
    sheet.cell(row=3, column=18).value = "AS A % OF TOTAL SHARES HELD"
    sheet.cell(row=3, column=20).value = "SHAREHOLDING (NO. OF SHARES) UNDER"
    sheet.cell(row=4, column=9).value = "CLASS X"
    sheet.cell(row=4, column=10).value = "CLASS Y"
    sheet.cell(row=4, column=11).value = "TOTAL"
    sheet.cell(row=4, column=15).value = "(A)"
    sheet.cell(row=4, column=16).value = "(B)"
    sheet.cell(row=4, column=17).value = "(A)"
    sheet.cell(row=4, column=18).value = "(B)"
    sheet.cell(row=4, column=20).value = "SUBCATEGORY (I)"
    sheet.cell(row=4, column=21).value = "SUBCATEGORY (II)"
    sheet.cell(row=4, column=22).value = "SUBCATEGORY (III)"
    
    # Define and apply custom style
    custom_style_name = "Blue_Header"
    custom_style = None
    
    for style in wb_obj.named_styles:
        if style == custom_style_name:
            custom_style = style

    # Add the custom style if it doesnt exist
    if custom_style is None:
        custom_style = NamedStyle(name=custom_style_name)
        custom_style.font = Font(name='Roboto', color='00FFFFFF')
        custom_style.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        custom_fill = PatternFill(start_color='FF3A2D7D', end_color='FF3A2D7D', fill_type='solid')
        custom_style.fill = custom_fill
        custom_style.border = Border(left=Side(style='thin', color='00FFFFFF'),
                                     right=Side(style='thin', color='00FFFFFF'),
                                     top=Side(style='thin', color='00FFFFFF'),
                                     bottom=Side(style='thin', color='00FFFFFF'))
        wb_obj.add_named_style(custom_style)
    
    # Apply custom style to the specified range of cells
    for row_idx in range(1, 5):
        for col_idx in range(1, 23):  # Adjust range according to your needs
            sheet.cell(row=row_idx, column=col_idx).style = custom_style